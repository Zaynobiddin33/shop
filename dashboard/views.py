from django.shortcuts import render, redirect
from django.contrib.auth.models import User
from main.models import *
from PIL import Image
from openpyxl import Workbook
from django.http import HttpResponse
from io import BytesIO
from openpyxl import load_workbook
from datetime import datetime
from string import ascii_letters
from .funcs import search, page_generate
from django.core.paginator import Paginator
from django.core.exceptions import FieldError


def dashboard(request):
    categorys =  Category.objects.all()
    products = Product.objects.all()
    users = User.objects.all()
    numusers = User.objects.all().count()
    numproducts = Product.objects.all().count()
    numcategories = Category.objects.all().count()
    context = {
        'categorys':categorys,
        'products':products,
        'users':users,
        'numusers':numusers,
        'numproducts':numproducts,
        'numcategories':numcategories
    }
    return render(request, 'dashboard/index.html', context)


def category_list(request):
    categories = Category.objects.all()
    p = Paginator(categories, 3)
    page = request.GET.get('page')
    categories = p.get_page(page)
    return render(request, 'dashboard/category/list.html', {'categories':categories})


def category_detail(request, id):
    category = Category.objects.get(id=id)
    products = Product.objects.filter(category=category, is_active=True)
    context = {
        'category':category,
        'products':products
    }
    return render(request, 'category/list.html', context)


def category_update(request, id):
    category = Category.objects.get(id=id)
    category.name = request.POST['name']
    category.save()

    return redirect('category_detail', {'category':category})


def category_delete(request, id):
    category = Category.objects.get(id=id)
    category.delete()
    return redirect('dashboard:category_list')

def category_create(request):
    if request.method == 'POST':
        Category.objects.create(
            name = request.POST['name']
        )
    return render(request, 'dashboard/category/create.html')



# products

# def products(request):
#     products = Product.objects.all()
#     category = Category.objects.all()
#     name = request.GET.get('name')
#     category_id = request.GET.get('category')
#     name = request.GET.get('name')
#     currency = request.GET.get('currency')
#     price = request.GET.get('price')
#     if name:
#         products = Product.objects.filter(name = name, price = price, category_id = category_id, currency = currency)

#     return render(request, 'dashboard/products/list.html',{'products':products, 'categories':category})


def product_create(request):
    categorys = Category.objects.all()
    context = {
        'categorys':categorys
    }
    if request.method == "POST":
        name = request.POST['name']
        description = request.POST['description']
        quantity = request.POST['quantity']
        price = request.POST['price']
        currency = request.POST['currency']
        baner_image = request.FILES['baner_image']
        category_id = request.POST['category_id']
        images = request.FILES.getlist('images')
        product = Product.objects.create(
            name=name,
            description = description,
            quantity=quantity,
            price=price,
            currency=currency,
            baner_image=baner_image,
            category_id=category_id
        )
        Overall.objects.create(
            product = product
        )
        for image in images:
            ProductImage.objects.create(
                image=image,
                product=product
            
            )

    return render(request, 'dashboard/products/create.html', context)

def product_update(request, id):
    product = Product.objects.get(id = id)
    images = ProductImage.objects.filter(product = product)
    context = {
        'product':product,
        'categorys':Category.objects.all(),
        'images': images
    }
    if request.method == "POST":
        product.name = request.POST['name']
        product.description = request.POST['description']
        product.quantity = request.POST['quantity']
        product.price = request.POST['price']
        product.currency = request.POST['currency']
        try:
            product.baner_image = request.FILES['baner_image']
        except:
            pass
        product.category_id = request.POST['category_id']
        images = request.FILES.getlist('images')
        for image in images:
            ProductImage.objects.create(
                image=image,
                product=product
            
            )
        product.save()
    return render (request, 'dashboard/products/update.html', context)

def product_delete(request, id):
    Product.objects.get(id = id).delete()
    return redirect('dashboard:products')

def image_delete(request, id):
    ProductImage.objects.get(id = id).delete()
    previous_url = request.META.get('HTTP_REFERER')
    return redirect(previous_url )


def add_admin(request):
    users = User.objects.all()
    context = {'users':users}
    if request.method == "POST":
        if User.objects.filter(username = request.POST['user']).first():
            username = request.POST['user']
            new_admin = User.objects.get(username = username)
            new_admin.is_staff = True
            new_admin.save()
    return render(request, 'dashboard/user/add.html', context)

def admins(request):
    admins = User.objects.filter(is_staff = True)
    return render(request, 'dashboard/user/list.html', {'admins': admins})

def delete_admin(request, id):
    User.objects.get(id = id).delete()
    return redirect('dashboard:admins')

#creating income
def income(request):
    products = Product.objects.all()

    if request.method == "POST":
        product = Product.objects.get(id = request.POST['product_id'])
        quantity= request.POST['quantity']
        ProductIncome.objects.create(
            product = product,
            amount = int(quantity)
        )
        product.quantity+=int(request.POST['quantity'])
        product.save()
        return redirect ('dashboard:income')
    return render(request, 'dashboard/income/add.html', {"products": products})

def list_income(request):
    incomes = ProductIncome.objects.all().order_by('-date')
    p = Paginator(incomes, 10)
    page = request.GET.get('page')
    incomes = p.get_page(page)
    context = {'enters':incomes}
    
    return render(request, 'dashboard/income/list.html', context)

def delete_income(request, id):
    ProductIncome.objects.get(id=id).delete()
    return redirect('dashboard:list_enter')

def update_income(request, id):
    if request.method == 'POST':
        quantity = int(request.POST['quantity'])
        income = ProductIncome.objects.get(id=id)
        income.amount = quantity
        income.save()
    return redirect('dashboard:list_enter')

#Excel generator
def income_excel(request):
    objs = ProductIncome.objects.all().order_by('-date')
    wb = Workbook()
    wsh = wb.active
    headers = ['No','Product name', 'Product amount', 'Added date']
    wsh.append(headers)

    for i, obj in enumerate(objs):
        row_data = [i+1, obj.product.name, obj.amount, str(obj.date)]
        wsh.append(row_data)

    for col in wsh.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        changed_width = (max_length + 2) * 1.2
        wsh.column_dimensions[column].width = changed_width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="incomes.xlsx"'
    return response

#excel for 'kirim va chiqimlar' page
def overall_excel(request):
    objs = Overall.objects.all().order_by('-id')
    wb = Workbook()
    wsh = wb.active
    headers = ['No','Product name', 'Income', 'Outcome']
    wsh.append(headers)

    for i, obj in enumerate(objs):
        row_data = [i+1,obj.product.name, obj.all_income, obj.all_outcome]
        wsh.append(row_data)

    for col in wsh.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        changed_width = (max_length + 2) * 1.2
        wsh.column_dimensions[column].width = changed_width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="incomes.xlsx"'
    return response


def income_outcome(request):
    data = Overall.objects.all()
    p = Paginator(data, 10)
    page = request.GET.get('page')
    data = p.get_page(page)
    context = {'data': data}
    return render (request, 'dashboard/income/income_outcome.html', context)

def excel_input(request):
    if request.method == "POST":
        file = request.FILES['file']
        workbook = load_workbook(file)
        ws = workbook.active
        for row in ws.iter_rows(values_only=True):
            if str(row[1])!='None' and str(row[2])!='None' and str(row[3])!='None' and row[3][0] not in ascii_letters:
                product = Product.objects.filter(name = row[1]).first()
                date = datetime.strptime(row[3], "%Y-%m-%d %H:%M:%S.%f%z")
                if product:
                    ProductIncome.objects.create(
                        product = product,
                        amount = int(row[2]),
                        date = date
                    )
    return render (request, 'dashboard/excel/input.html')

def product_detail(request, id):
    product = Product.objects.get(id = id)
    images = ProductImage.objects.filter(product_id = id)
    all = Product_income_outcome.objects.filter(product = product).order_by("-date")
    p = Paginator(all, 10)
    page = request.GET.get('page')
    all = p.get_page(page)
    context = {'product': product,
               'images': images,
               'all': all}
    
    return render (request, 'dashboard/detail/detail.html', context)

def products(request):
    result = search(request)
    categories  = Category.objects.all() 
    enters = Product.objects.filter(**result)
    p = Paginator(enters, 10)
    page = request.GET.get('page')
    enters = p.get_page(page)
    context = {'products': enters, 'categories':categories}
    return render(request, 'dashboard/products/list.html', context)
